import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from dlt_sources import create_medicloud_connection

def normalize_code(value):
    """Normalize procedure codes for robust matching: lowercase, trim, remove internal spaces."""
    if value is None:
        return ""
    try:
        return str(value).strip().lower().replace(" ", "")
    except Exception:
        return str(value)

def load_and_clean_data(first_file):
    """
    Load and clean the tariff data from MediCloud and FIRST data from CSV
    """
    try:
        # Connect to MediCloud and load tariff data
        conn = create_medicloud_connection()
        
        # Load tariff data
        tariff_query = """
            SELECT 
                trf.tariffid,
                trf.tariffname,
                trfp.tariffamount,
                trfp.procedurecode,
                trf.expirydate
            FROM dbo.tariff trf
            JOIN dbo.procedure_tariff trfp ON trf.tariffid = trfp.tariffid
            WHERE CAST(trf.expirydate AS DATETIME) >= CAST(GETDATE() AS DATETIME)
            AND trfp.procedurecode IS NOT NULL
            AND trfp.tariffamount > 0
        """
        tariff_df = pd.read_sql(tariff_query, conn)
        print(f"Successfully loaded {len(tariff_df)} tariff records from MediCloud")
        
        # Load claims data for frequency calculation
        claims_query = """
            SELECT 
                procedurecode,
                COUNT(DISTINCT claimid) as frequency
            FROM dbo.claims
            WHERE datesubmitted >= DATEADD(year, -1, GETDATE())
            AND procedurecode IS NOT NULL
            AND approvedamount > 0
            GROUP BY procedurecode
        """
        claims_df = pd.read_sql(claims_query, conn)
        conn.close()
        print(f"Successfully loaded {len(claims_df)} claims records from MediCloud")
        
    except Exception as e:
        print(f"Error loading from MediCloud: {str(e)}")
        raise
    
    # Load FIRST data from CSV
    try:
        first_df = pd.read_csv(first_file)
        print(f"Successfully loaded {len(first_df)} records from CSV")
    except Exception as e:
        print(f"Error loading FIRST data from CSV: {str(e)}")
        raise
    
    # Clean column names (remove extra spaces)
    tariff_df.columns = tariff_df.columns.str.strip()
    first_df.columns = first_df.columns.str.strip()
    
    # Ensure procedurecode columns are strings for proper matching
    tariff_df['procedurecode'] = tariff_df['procedurecode'].astype(str)
    first_df['procedurecode'] = first_df['procedurecode'].astype(str)
    claims_df['procedurecode'] = claims_df['procedurecode'].astype(str)
    
    # Normalize procedure codes for robust matching
    tariff_df['procedurecode_norm'] = tariff_df['procedurecode'].apply(normalize_code)
    first_df['procedurecode_norm'] = first_df['procedurecode'].apply(normalize_code)
    claims_df['procedurecode_norm'] = claims_df['procedurecode'].apply(normalize_code)
    
    # Ensure numeric columns are properly formatted
    tariff_df['tariffamount'] = pd.to_numeric(tariff_df['tariffamount'], errors='coerce')
    first_df['price'] = pd.to_numeric(first_df['price'], errors='coerce')
    claims_df['frequency'] = pd.to_numeric(claims_df['frequency'], errors='coerce')
    
    # Merge frequency data from claims into first_df
    first_df = first_df.merge(claims_df[['procedurecode_norm', 'frequency']], 
                             on='procedurecode_norm', 
                             how='left')
    first_df['frequency'] = first_df['frequency'].fillna(0).astype(int)
    
    print(f"Successfully merged frequency data from claims")
    
    return tariff_df, first_df

def analyze_tariffs(tariff_df, first_df, output_file='tariff_analysis_results.xlsx'):
    """
    Analyze tariffs against FIRST data and create Excel sheets for each procedure code
    """
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Get unique procedure codes from FIRST data (using normalized codes)
    unique_procedures = first_df['procedurecode_norm'].unique()
    
    # Create summary data for overview
    summary_data = []
    
    print(f"Processing {len(unique_procedures)} procedure codes...")
    
    for i, proc_code_norm in enumerate(unique_procedures):
        # Get original procedure code for display
        original_code = first_df[first_df['procedurecode_norm'] == proc_code_norm]['procedurecode'].iloc[0]
        print(f"Processing {i+1}/{len(unique_procedures)}: {original_code}")
        
        # Get price value and description for this procedure code from FIRST data
        first_row = first_df[first_df['procedurecode_norm'] == proc_code_norm]
        
        if first_row.empty:
            continue
            
        price_amount = first_row['price'].iloc[0]
        description = first_row.get('description', 'No description available').iloc[0] if 'description' in first_row.columns else 'No description available'
        
        # Skip if price_amount is NaN
        if pd.isna(price_amount):
            continue
        
        # Filter tariff data for this procedure code where tariffamount > price
        filtered_tariffs = tariff_df[
            (tariff_df['procedurecode_norm'] == proc_code_norm) & 
            (tariff_df['tariffamount'] > price_amount)
        ]
        
        # Create sheet name (Excel sheet names have character limits)
        sheet_name = str(original_code)[:31]  # Excel sheet name limit is 31 characters
        
        # If there are results, create a sheet
        if not filtered_tariffs.empty:
            # Select relevant columns
            result_df = filtered_tariffs[['tariffname', 'tariffamount', 'procedurecode']].copy()
            
            # Add price amount from FIRST for reference
            result_df['price_from_first'] = price_amount
            result_df['difference'] = result_df['tariffamount'] - price_amount
            
            # Sort by tariffamount descending
            result_df = result_df.sort_values('tariffamount', ascending=False)
            
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add header information
            ws['A1'] = f"Procedure Code: {original_code}"
            ws['A2'] = f"Description: {description}"
            ws['A3'] = f"Price from FIRST: {price_amount:,.2f}"
            ws['A4'] = f"Number of tariffs exceeding price: {len(result_df)}"
            ws['A5'] = f"Frequency from claims: {first_row['frequency'].iloc[0]}"
            ws['A6'] = ""  # Empty row
            
            # Add the dataframe starting from row 7
            for r in dataframe_to_rows(result_df, index=False, header=True):
                ws.append(r)
            
            # Add to summary
            summary_data.append({
                'procedurecode': original_code,
                'price_from_first': price_amount,
                'frequency': first_row['frequency'].iloc[0],
                'count_exceeding': len(result_df),
                'highest_tariff': result_df['tariffamount'].max(),
                'total_procedures_in_tariff': len(tariff_df[tariff_df['procedurecode_norm'] == proc_code_norm])
            })
        else:
            # Add to summary even if no exceeding tariffs found
            total_procedures = len(tariff_df[tariff_df['procedurecode_norm'] == proc_code_norm])
            summary_data.append({
                'procedurecode': original_code,
                'price_from_first': price_amount,
                'frequency': first_row['frequency'].iloc[0],
                'count_exceeding': 0,
                'highest_tariff': 0,
                'total_procedures_in_tariff': total_procedures
            })
    
    # Create summary sheet
    summary_df = pd.DataFrame(summary_data)
    summary_df = summary_df.sort_values('count_exceeding', ascending=False)
    
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws['A1'] = "TARIFF ANALYSIS SUMMARY"
    summary_ws['A2'] = f"Total Procedure Codes Analyzed: {len(unique_procedures)}"
    summary_ws['A3'] = f"Codes with Exceeding Tariffs: {len(summary_df[summary_df['count_exceeding'] > 0])}"
    summary_ws['A4'] = f"Total Frequency from Claims: {summary_df['frequency'].sum()}"
    summary_ws['A5'] = ""
    
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        summary_ws.append(r)
    
    # Save the workbook
    wb.save(output_file)
    print(f"\nAnalysis complete! Results saved to: {output_file}")
    print(f"Created {len([sheet for sheet in wb.sheetnames if sheet != 'Summary'])} procedure code sheets")
    
    return summary_df

def main():
    """
    Main function to run the analysis
    """
    # File paths - UPDATE THESE WITH YOUR ACTUAL FILE PATHS
    first_file = "second.csv"
    output_file = "tariff_analysis_results.xlsx"
    
    try:
        # Check if FIRST file exists
        if not os.path.exists(first_file):
            print(f"Error: FIRST CSV file '{first_file}' not found!")
            return
        
        print("Loading data from MediCloud and FIRST file...")
        tariff_df, first_df = load_and_clean_data(first_file)
        
        print(f"Loaded {len(tariff_df)} tariff records and {len(first_df)} FIRST records")
        
        # Run analysis
        summary = analyze_tariffs(tariff_df, first_df, output_file)
        
        # Print summary statistics
        print("\n" + "="*50)
        print("ANALYSIS SUMMARY")
        print("="*50)
        print(f"Total procedure codes: {len(summary)}")
        print(f"Codes with exceeding tariffs: {len(summary[summary['count_exceeding'] > 0])}")
        print(f"Total exceeding tariffs found: {summary['count_exceeding'].sum()}")
        print("\nTop 5 procedure codes with most exceeding tariffs:")
        top_5 = summary.head(5)[['procedurecode', 'count_exceeding', 'price_from_first', 'frequency']]
        print(top_5.to_string(index=False))
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()

# Alternative function for when you have the dataframes already loaded or custom paths
def analyze_with_custom_paths(first_csv_path, output_file='tariff_analysis_results.xlsx'):
    """
    Use this function with custom file paths
    
    Parameters:
    first_csv_path: Path to the FIRST CSV file
    output_file: Name of the output Excel file
    """
    tariff_df, first_df = load_and_clean_data(first_csv_path)
    return analyze_tariffs(tariff_df, first_df, output_file)

# Alternative function for when you have the dataframes already loaded
def analyze_loaded_dataframes(tariff_df, first_df, output_file='tariff_analysis_results.xlsx'):
    """
    Use this function if you already have the dataframes loaded in memory
    
    Parameters:
    tariff_df: DataFrame with columns ['tariffid', 'tariffname', 'tariffamount', 'procedurecode', 'expirydate']
    first_df: DataFrame with columns ['procedurecode', 'price', 'frequency']
    output_file: Name of the output Excel file
    """
    return analyze_tariffs(tariff_df, first_df, output_file)